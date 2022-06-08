import numpy as np
import pandas as pd
import cv2
import pytesseract
import imutils
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from PIL import ImageTk, Image
import openpyxl

pytesseract.pytesseract.tesseract_cmd = r"D:\Program Files\Tesseract-OCR\tesseract.exe"

root = Tk()
root.title("Prepoznavanje tablica automobila i automatska naplata parkinga na osnovu slike tablice")
root.geometry("1440x1080")

act_text = ""

def open():
    global car_image
    root.filename = filedialog.askopenfilename(initialdir="D:\Program Files\Python\ANPR_Projekt\Tablice_Baza", title="Odaberite fotografiju", filetypes=(("jpg files","*.jpg"),("png files", "*.png"),("all files", "*.*")))    
    car_image = ImageTk.PhotoImage(Image.open(root.filename))    
    car_image_label = Label(image=car_image).grid(row=1,column=0)

def check():
    image = cv2.imread(root.filename)
    image = imutils.resize(image, width=1080)
    cv2.imshow("Original", image)

    # izrezana slika 
    cropped_image = image[900:1180, 250:880]
    cv2.imshow("Cropped", cropped_image)   

    # pretvaranje slike u sliku sivih tonova
    gray = cv2.cvtColor(cropped_image, cv2.COLOR_BGR2GRAY)
    cv2.imshow("Grayscale Conversion", gray)

    # slika se blura ili zamucuje da bi se smanjio sum
    gray = cv2.bilateralFilter(gray, 11, 17, 50)
    cv2.imshow("Bilateral Filter", gray)

    # u ovom dijelu se vrsi detekcija rubova
    edged = cv2.Canny(gray, 100, 200)
    cv2.imshow("Canny Edges", edged)

    cv2.waitKey(0)
    cv2.destroyAllWindows()

    # pronalazenje rubova na "edged" slici 
    (cnts, _) = cv2.findContours(edged.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    cnts=sorted(cnts, key = cv2.contourArea, reverse = True)[:30]

    NumberPlateCnt = None
    count = 0

    for c in cnts:
        
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.02 * peri, True)
        # ako aproksimirana kontura ima četiri točke, pretpostavlja se da je zaslon pronađen
        if len(approx) == 4:
            NumberPlateCnt = approx
            break

    # maskiraju se ostatak slike tj. sve osim reg. plocice
    mask = np.zeros(gray.shape,np.uint8)
    new_image = cv2.drawContours(mask,[NumberPlateCnt],0,255,-1)
    new_image = cv2.bitwise_and(cropped_image,cropped_image,mask=mask)
    cv2.namedWindow("Samo tablica",cv2.WINDOW_NORMAL)
    cv2.imshow("Samo tablica",new_image)

    ray = cv2.cvtColor(new_image, cv2.COLOR_BGR2GRAY)
    cv2.imshow("Grayscale samo tablice", ray)

    ret,thresh1 = cv2.threshold(ray, 100, 255, cv2.THRESH_BINARY)
    cv2.imshow("Binarna slika samo tablice", thresh1)

    # tesseract conf
    config = ('-l eng --oem 1 --psm 3')    

    # pokrece se tesseract OCR na slici
    act_text = pytesseract.image_to_string(thresh1, config=config)        

    cv2.waitKey(0)
    cv2.destroyAllWindows() 

    pay(act_text)       

def pay(lic_str):

    c = 0

    wrkbk = openpyxl.load_workbook("UsersDB.xlsx")
    
    sh = wrkbk.active    
   
    for i in range(2, sh.max_row+1):              
        for j in range(1, sh.max_column+1):   
            cell_obj = sh.cell(row=i, column=j)        
            if ("{}".format(sh.cell(row=i, column=j).value) + "\n") == lic_str: 
                c += 1               
                sh.cell(row=i, column=j+1).value = sh.cell(row=i, column=j+1).value - 4    
                messagebox.showinfo("Obavijest","Korisniku registarskih tablica " + lic_str + " je naplacen parking u iznosu od 4 HRK.")                   
                         
    if c == 0:
        messagebox.showerror("Greška","Korisnik nije pronađen")                         

    wrkbk.save(filename="UsersDB.xlsx")       


open_btn = Button(root, text="Odabir fotografije", command=open).grid(row=0,column=0)
check_btn = Button(root, text="Provjeri",command=check).grid(row=0,column=1)

root.mainloop()
